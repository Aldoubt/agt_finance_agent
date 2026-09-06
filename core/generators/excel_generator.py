"""Generate the procurement summary workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelGenerator:
    HEADERS = [
        "序号",
        "商品名称",
        "金额",
        "发票是否存在",
        "实物图是否存在",
        "购买凭证是否存在",
        "支付凭证是否存在",
        "发票文件",
    ]

    def generate(self, items, output_path: str | Path, total_amount: float) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "采购统计"
        ws.append(self.HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for index, item in enumerate(items, 1):
            ws.append(
                [
                    index,
                    item.name,
                    item.gross_amount,
                    "是" if item.invoice_files else "否",
                    "是" if item.photo_files else "否",
                    "是" if item.purchase_proof_files else "否",
                    "是" if item.payment_files else "否",
                    ",".join(Path(x).name for x in item.invoice_files),
                ]
            )

        ws.append([])
        ws.append(["总金额", "", round(float(total_amount or 0), 2)])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.cell(ws.max_row, 3).font = Font(bold=True)

        widths = [10, 42, 14, 16, 18, 20, 20, 35]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = width

        wb.save(output)
        return output

