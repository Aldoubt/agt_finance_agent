"""Build a machine-readable ground truth snapshot from manual final outputs."""

import json
from pathlib import Path

import openpyxl

from .docx_truth_loader import load_photo_truth


def _excel_items(path: str | Path) -> list[dict]:
    sheet = openpyxl.load_workbook(path, data_only=True).active
    headers = [sheet.cell(2, col).value for col in range(1, sheet.max_column + 1)]

    merged_values = {}
    for merged in sheet.merged_cells.ranges:
        value = sheet.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                merged_values[(row, col)] = value

    def cell_value(row: int, col: int):
        value = sheet.cell(row, col).value
        if value is not None:
            return value
        return merged_values.get((row, col))

    rows = []
    for row_index in range(3, sheet.max_row + 1):
        serial = sheet.cell(row_index, 2).value
        name = sheet.cell(row_index, 3).value
        if not isinstance(serial, (int, float)) or not name:
            continue
        row = {
            headers[col - 1]: cell_value(row_index, col)
            for col in range(1, sheet.max_column + 1)
            if headers[col - 1]
        }
        rows.append(row)
    return rows


def build(
    truth_excel: str | Path,
    truth_docx: str | Path,
    source_image_dir: str | Path,
    output_json: str | Path,
) -> dict:
    photo_truth = load_photo_truth(truth_docx, source_image_dir)
    result = {
        "items": _excel_items(truth_excel),
        "photo_truth": [x.to_dict() for x in photo_truth],
    }
    target = Path(output_json)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return result


if __name__ == "__main__":
    import sys

    result = build(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
    matched = sum(1 for x in result["photo_truth"] if x["source_image"])
    print(json.dumps({"item_count": len(result["items"]), "photo_truth_count": len(result["photo_truth"]), "matched_source_images": matched}, ensure_ascii=False))

